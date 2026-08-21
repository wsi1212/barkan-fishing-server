#!/usr/bin/env python3
"""
gen_balance_workbook.py — 바르칸 열도 「밸런스 작업용 통합 엑셀」 생성기.

외부 밸런서(사람)에게 넘길 단일 .xlsx 를 라이브 권위 소스에서 **매번 다시 뽑는다**.
사본을 고정해 두면 원본 갱신에 안 따라가므로(프로젝트 규칙) 손으로 만든 시트는 두지 않는다.

권위 소스 3종
  1. JSON  : plugins/BlockShip/*.json            (부품·물고기·재료·레시피·강화·퀘스트·지역·환경)
  2. Java  : blockship-plugin src 하드코딩 상수   (등급 PRD·판매공식·레벨곡선·특성트리)
  3. stats.db catalog_version                     (요리·통발·작물·드릴·섬/길드가격 — 코드 상수를
             플러그인이 부팅마다 정규화 JSON 으로 스냅샷한 것. Java 정규식 파싱보다 안전)

사용법
    python3 gen_balance_workbook.py --out ~/Desktop/바르칸_밸런스.xlsx
    python3 gen_balance_workbook.py --json <BlockShip디렉터리> --sim 400000

openpyxl 필요:  python3 -m pip install --user openpyxl
"""

import argparse
import datetime
import gzip
import json
import os
import random
import re
import sqlite3
import sys
from collections import Counter, OrderedDict, defaultdict

JAVA_ROOT = os.environ.get(
    "BLOCKSHIP_JAVA",
    "/Users/user/development/blockship-plugin/src/main/java/com/blockship",
)
JSON_ROOT = os.environ.get(
    "BLOCKSHIP_JSON",
    "/Users/user/Library/Application Support/feather/player-server/servers/"
    "07de2d81-991a-47e2-b62d-06c0d1b5150a/plugins/BlockShip",
)

WARNINGS = []


def warn(msg):
    WARNINGS.append(msg)
    print(f"  !! {msg}", file=sys.stderr)


# ─────────────────────────────────────────────────────────── 소스 로더
def read_java(rel):
    try:
        with open(os.path.join(JAVA_ROOT, rel), encoding="utf-8") as f:
            return f.read()
    except OSError as e:
        warn(f"Java 못 읽음: {rel} ({e})")
        return ""


def read_json(name, default=None):
    try:
        with open(os.path.join(JSON_ROOT, name), encoding="utf-8") as f:
            return json.load(f)
    except (OSError, json.JSONDecodeError) as e:
        warn(f"JSON 못 읽음: {name} ({e})")
        return default


def read_catalogs():
    """stats.db catalog_version 에서 코드 상수 스냅샷(kind→dict)을 뽑는다."""
    out, meta = {}, {}
    path = os.path.join(JSON_ROOT, "telemetry", "stats.db")
    if not os.path.exists(path):
        warn(f"stats.db 없음 ({path}) — 요리/통발/작물/드릴/섬가격 시트가 비게 된다")
        return out, meta
    try:
        con = sqlite3.connect(f"file:{path}?mode=ro", uri=True)
        for (kind,) in con.execute("SELECT DISTINCT kind FROM catalog_version"):
            ts, blob = list(
                con.execute(
                    "SELECT ts, json_gz FROM catalog_version WHERE kind=? ORDER BY ts DESC LIMIT 1",
                    (kind,),
                )
            )[0]
            out[kind] = json.loads(gzip.decompress(blob))
            meta[kind] = datetime.datetime.fromtimestamp(ts / 1000.0)
        con.close()
    except Exception as e:  # noqa: BLE001
        warn(f"stats.db 읽기 실패: {e}")
    return out, meta


def java_mention_index():
    """모든 Java 소스를 한 번 읽어 '어느 파일이 이 문자열을 언급하는가' 검색기를 만든다.

    재료 획득 경로가 JSON 에만 있는 게 아니라 코드에 있는 경우(RecipeLoader 가 주입하는 레시피,
    DrillManager 채굴 산출, ImugiBattle 보스 드롭 등)를 놓치지 않기 위한 것.
    """
    files = []
    for root, _dirs, names in os.walk(JAVA_ROOT):
        for n in names:
            if not n.endswith(".java"):
                continue
            p = os.path.join(root, n)
            try:
                with open(p, encoding="utf-8") as f:
                    files.append((os.path.relpath(p, JAVA_ROOT), f.read()))
            except OSError:
                continue
    if not files:
        warn("Java 소스를 하나도 못 읽었다 — 재료 획득경로 분류가 부실해진다")

    # 파일 경로 → 사람이 읽을 획득 경로 라벨
    LABEL = [
        ("drill/", "드릴 채굴"),
        ("boss/", "보스 드롭"),
        ("crafting/RecipeLoader", "코드 주입 레시피"),
        ("crop/", "특수작물 수확"),
        ("forage/", "채집"),
        ("trap/", "통발"),
        ("fishing/WetTreasureChest", "젖은 보물상자"),
        ("quest/", "퀘스트 보상"),
        ("economy/", "상점"),
        ("cooking/", "요리 재료"),
    ]

    def where(token):
        hits, labels = [], []
        for rel, text in files:
            if token in text:
                hits.append(rel)
                for pfx, lab in LABEL:
                    if rel.startswith(pfx) and lab not in labels:
                        labels.append(lab)
        return hits, labels

    return where


def jnums(text):
    return [
        float(x) if ("." in x or "e" in x.lower()) else int(x)
        for x in re.findall(r"-?\d[\d_]*\.?\d*(?:[eE]-?\d+)?", text.replace("_", ""))
    ]


# ─────────────────────────────────────────────────────────── Java 상수
def pull_java_consts():
    c = {}

    gr = read_java("fishing/GradeRoller.java")
    m = re.search(r"ROLL_ORDER\s*=\s*\{(.*?)\};", gr, re.S)
    rolls = []
    if m:
        for g, base, gate in re.findall(
            r'RollEntry\(\s*"(\w+)"\s*,\s*([\d.]+)\s*,\s*(\d+)', m.group(1)
        ):
            rolls.append((g, float(base), int(gate)))
    if not rolls:
        warn("GradeRoller.ROLL_ORDER 파싱 실패 — 등급 확률 시트 확인 필요")
    c["roll_order"] = rolls
    c["grade_unlock"] = [
        (int(l), int(g)) for l, g in re.findall(r"level\s*>=\s*(\d+)\)\s*m\s*=\s*(\d+)", gr)
    ]

    fi = read_java("economy/FishItem.java")
    m = re.search(r"long base = switch \(grade\) \{(.*?)\};", fi, re.S)
    prices = {}
    if m:
        for g, v in re.findall(r'case "(\w)" -> (\d+)', m.group(1)):
            prices[g] = int(v)
    if not prices:
        warn("FishItem.fishPrice 등급 기본가 파싱 실패")
    c["grade_price"] = prices
    fresh = re.findall(r"if \(ageMins <= (\d+)\) return ([\d.]+);", fi)
    c["freshness"] = [(int(a), float(b)) for a, b in fresh]
    m = re.search(r"return ([\d.]+);\s*\n\s*\}\s*\n\s*/\*\* 등급 기본가", fi)
    c["freshness_tail"] = float(m.group(1)) if m else 0.20

    rm = read_java("fishing/RewardMath.java")
    m = re.search(r"int baseExp\(String grade\) \{\s*return switch \(grade\) \{(.*?)\};", rm, re.S)
    bexp = {}
    if m:
        for g, v in re.findall(r'case "(\w)" -> (\d+)', m.group(1)):
            bexp[g] = int(v)
        m2 = re.search(r"default -> (\d+);", m.group(1))
        if m2:
            bexp["E"] = int(m2.group(1))
    if not bexp:
        warn("RewardMath.baseExp 파싱 실패")
    c["base_exp"] = bexp
    m = re.search(r"LevelBonus levelBonus\(int lv\) \{(.*?)return new LevelBonus", rm, re.S)
    c["level_bonus_src"] = m.group(1) if m else ""

    lm = read_java("fishing/FishingLevelManager.java")
    m = re.search(r"NEED_TABLE\s*=\s*new\s+int\[\]\s*\{(.*?)\}", lm, re.S)
    c["need_table"] = [int(x) for x in re.findall(r"\d+", m.group(1))] if m else []
    if not c["need_table"]:
        warn("NEED_TABLE 파싱 실패 — 레벨곡선 시트 비게 됨")

    return c


def parse_level_bonus(src):
    """levelBonus() 계단을 (레벨, 항목, 증가량) 목록으로."""
    rows = []
    for lv, body in re.findall(r"if \(lv >= (\d+)\)\s*\{?([^\n]*)", src):
        lv = int(lv)
        for key, field in (("crit +=", "크리확률"), ("critDmg +=", "크리배율"), ("escRed +=", "도주감소")):
            m = re.search(re.escape(key) + r"\s*(\d+)", body)
            if m:
                rows.append((lv, field, int(m.group(1))))
        m = re.search(r'maxGrade = "(\w)"', body)
        if m:
            rows.append((lv, "최대등급 해금", m.group(1)))
    rows.sort(key=lambda r: (r[0], str(r[1])))
    return rows


def parse_skill_tree():
    src = read_java("skilltree/SkillTreeManager.java")
    trees = []
    for tm in re.finditer(
        r'return new SkillTree\("([^"]+)",\s*"([^"]+)",\s*"([^"]+)"', src
    ):
        trees.append((tm.group(1), tm.group(2), tm.group(3), tm.start()))
    rows = []
    # 각 트리 메서드 블록 안의 Node.stat / Node.proc 을 순서대로 수집
    blocks = re.split(r"private static SkillTree \w+\(\)", src)
    names = re.findall(r"private static SkillTree (\w+)\(\)", src)
    for name, block in zip(names, blocks[1:]):
        tm = re.search(r'return new SkillTree\("([^"]+)",\s*"([^"]+)"', block)
        skill = tm.group(1) if tm else name
        for m in re.finditer(
            r'Node\.stat\(\s*"([^"]*)",\s*"([^"]*)",\s*"([^"]*)",\s*Material\.\w+,\s*(\d+),\s*'
            r'(?:"([^"]*)"|null),\s*(\d+),\s*(\d+),\s*(?:"([^"]*)"|null),\s*([\d.]+),\s*"([^"]*)"'
            r',\s*"([^"]*)"',
            block,
        ):
            rows.append(
                dict(
                    숙련=skill, 계열=m.group(2), 노드id=m.group(1), 노드명=m.group(3),
                    종류="스탯", 최대랭크=int(m.group(4)), 선행=m.group(5) or "",
                    계열요구=int(m.group(6)), 스탯키=m.group(8) or "", 랭크당=float(m.group(9)),
                    단위=m.group(10), 최대효과=round(float(m.group(9)) * int(m.group(4)), 2),
                    효과설명=re.sub(r"§.", "", m.group(11)),
                )
            )
        for m in re.finditer(
            r'Node\.proc\(\s*"([^"]*)",\s*"([^"]*)",\s*"([^"]*)",\s*Material\.\w+,\s*'
            r'(?:"([^"]*)"|null),\s*(\d+),\s*(\d+),\s*(.*?)\)\s*[,;]',
            block,
            re.S,
        ):
            lines = " / ".join(re.sub(r"§.", "", s) for s in re.findall(r'"([^"]*)"', m.group(7)))
            rows.append(
                dict(
                    숙련=skill, 계열=m.group(2), 노드id=m.group(1), 노드명=m.group(3),
                    종류="발동(proc)", 최대랭크=1, 선행=m.group(4) or "",
                    계열요구=int(m.group(5)), 스탯키="", 랭크당="", 단위="", 최대효과="",
                    효과설명=lines,
                )
            )
    if not rows:
        warn("SkillTreeManager 노드 파싱 실패 — 특성 시트 비게 됨")
    return rows


# ─────────────────────────────────────────────────────────── 부품 파싱
STAT_ORDER = [
    "행운", "난이도", "등급업", "등급특화", "크기", "크리확률", "크리배율", "경험치",
    "더블찬스", "트리플찬스", "판매보너스", "도망감소", "내구보존",
    "수중호흡", "수영속도", "공격력", "공격속도", "돌진쿨감", "야간투시",
]
GRADE_RANK = {g: i for i, g in enumerate(["E", "D", "C", "B", "A", "S", "M", "L", "G"], 1)}


def parse_part(spec):
    """'이름|등급|가격|내구|스탯|레벨제한|출처' → dict"""
    f = spec.split("|")
    while len(f) < 7:
        f.append("")
    stats = {}
    for tok in f[4].split(","):
        tok = tok.strip()
        if not tok or ":" not in tok:
            continue
        k, v = tok.split(":", 1)
        try:
            stats[k.strip()] = float(v)
        except ValueError:
            stats[k.strip()] = v
    return dict(
        이름=f[0],
        등급=f[1],
        가격=int(f[2] or 0),
        내구=int(f[3] or 0),
        스탯=stats,
        레벨제한=int(f[5] or 0),
        출처=f[6],
    )


# ─────────────────────────────────────────────────────────── 등급 PRD 시뮬
_SIM_CACHE = {}


def sim_grade_rates(avail, casts, level=100, luck=0, roll_order=None, seed=20260822):
    """PRD(피티) 롤을 그대로 복제해 등급별 실제 출현율(%)을 낸다. avail=풀에 실제로 있는 등급 집합.

    ★가용성 폴백까지 복제한다(GradeRoller.roll 말미): 그 지역에 E 어종이 없으면 롤 실패분(=E)이
      전부 D→G 순 첫 가용 등급으로 승격된다. E가 없는 지역이 조용히 '최저등급 인플레'가 되는 이유.
    """
    key = (frozenset(avail), casts, level, luck)
    if key in _SIM_CACHE:
        return _SIM_CACHE[key]
    if not roll_order:
        return {}
    maxg = 6 + (level >= 30) + (level >= 45) + (level >= 60)
    luck_mult = (100.0 + luck) / 100.0
    pity = {g: 0 for g in "GLMSABCD"}
    counts = Counter()
    rnd = random.Random(seed)
    order = [(g, b, gate) for g, b, gate in roll_order]
    # 가용성 폴백 대상 = E 가 없을 때 실패분이 흘러갈 최저 가용 등급
    fallback = next((g for g in "DCBASMLG" if g in avail), None) if "E" not in avail else "E"
    for _ in range(casts):
        grade = "E"
        for g, base, gate in order:
            if grade != "E":
                break
            if gate > 0 and maxg < gate:
                continue
            if g not in avail:
                continue
            if rnd.random() < base * luck_mult * (1 + pity[g]) / 100.0:
                grade = g
                pity[g] = 0
        for k in pity:
            if k != grade:
                pity[k] += 1
        if grade == "E" and fallback and fallback != "E":
            grade = fallback   # E 어종이 없는 지역: 실패분이 최저 가용 등급으로 승격
        counts[grade] += 1
    res = {g: counts.get(g, 0) * 100.0 / casts for g in "EDCBASMLG"}
    _SIM_CACHE[key] = res
    return res


# ─────────────────────────────────────────────────────────── 엑셀 유틸
def build_workbook(out_path, sim_casts):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    HDR_FILL = PatternFill("solid", fgColor="1F3864")
    HDR_FONT = Font(color="FFFFFF", bold=True, size=10)
    SUB_FILL = PatternFill("solid", fgColor="D9E2F3")
    NOTE_FONT = Font(color="7F7F7F", size=9, italic=True)
    TITLE_FONT = Font(bold=True, size=13)
    THIN = Side(style="thin", color="BFBFBF")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    WARN_FILL = PatternFill("solid", fgColor="FFE699")
    BAD_FILL = PatternFill("solid", fgColor="F8CBAD")

    wb = Workbook()
    wb.remove(wb.active)
    index_rows = []

    def sheet(name, title, note, headers, rows, widths=None, tab=None, numfmt=None):
        ws = wb.create_sheet(name[:31])
        if tab:
            ws.sheet_properties.tabColor = tab
        ws["A1"] = title
        ws["A1"].font = TITLE_FONT
        ws["A2"] = note
        ws["A2"].font = NOTE_FONT
        hr = 4
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=hr, column=i, value=h)
            c.fill = HDR_FILL
            c.font = HDR_FONT
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = BORDER
        for r, row in enumerate(rows, hr + 1):
            for i, v in enumerate(row, 1):
                c = ws.cell(row=r, column=i, value=v)
                c.border = BORDER
                if numfmt and headers[i - 1] in numfmt:
                    c.number_format = numfmt[headers[i - 1]]
        ws.freeze_panes = f"A{hr + 1}"
        ws.auto_filter.ref = (
            f"A{hr}:{get_column_letter(len(headers))}{hr + max(len(rows), 1)}"
        )
        for i, h in enumerate(headers, 1):
            w = (widths or {}).get(h)
            if not w:
                sample = [len(str(h))] + [
                    len(str(r[i - 1])) for r in rows[:200] if i - 1 < len(r) and r[i - 1] is not None
                ]
                w = min(38, max(8, max(sample) + 2))
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.row_dimensions[hr].height = 30
        index_rows.append((name, title, len(rows)))
        return ws

    # ==================================================================== 로드
    print("소스 로딩...")
    P = read_json("parts.json", {"parts": {}, "order": []})
    F = read_json("fish.json", {"fish": {}, "regions": {}, "environment": {}})
    M = read_json("materials.json", {"materials": {}, "dropTables": {}})
    R = read_json("recipes.json", {"recipes": {}, "categories": {}})
    RG = read_json("regions.json", {})
    EN = read_json("enhance.json", {"order": [], "table": {}})
    EB = read_json("env-bonuses.json", {"weathers": {}, "times": {}})
    SV = read_json("submit-values.json", {})
    SH = read_json("shop-items.json", {"categories": []})
    QJ = read_json("quests.json", {})
    JC = pull_java_consts()
    CAT, CAT_TS = read_catalogs()

    parts = P.get("parts", {})
    fishdefs = F.get("fish", {})
    regionfish = F.get("regions", {})
    envfish = F.get("environment", {})
    mats = M.get("materials", {})
    drops = M.get("dropTables", {})
    recipes = R.get("recipes", {})
    roll_order = JC["roll_order"]
    gprice = JC["grade_price"]

    # 레시피 인덱스 ------------------------------------------------------
    rec_by_rod = {}
    rec_by_part = {}
    for rid, rc in recipes.items():
        if rc.get("rodPartName"):
            rec_by_rod[rc["rodPartName"]] = rc
        if rc.get("resultPartName"):
            rec_by_part[(rc.get("resultPartType", ""), rc["resultPartName"])] = rc

    def ing_text(rc, sep=" + "):
        if not rc:
            return ""
        return sep.join(
            f"{i.get('displayName') or i.get('typeOrMatId')} x{i.get('qty', 1)}"
            for i in rc.get("ingredients", [])
        )

    def ing_cols(rc, n=6):
        out = []
        ings = (rc or {}).get("ingredients", [])
        for i in range(n):
            if i < len(ings):
                out.append(f"{ings[i].get('displayName') or ings[i].get('typeOrMatId')}")
                out.append(ings[i].get("qty", 1))
            else:
                out += ["", None]
        return out

    # 지역 인덱스 --------------------------------------------------------
    def is_stub(rid):
        """실제 영역이 없는 껍데기 지역인가.

        ★shape=polygon/points3d 지역은 pos1/pos2 가 권위가 아니다(둘이 같아도 정상).
          예전 감사에서 늪지대/항구를 스텁으로 오판한 원인이 이것 — 다각형이 있으면 정상이다.
        """
        rd = RG.get(rid)
        if not rd:
            return True
        if rd.get("polygon") or rd.get("points3d"):
            return False
        p1, p2 = rd.get("pos1") or [0, 0, 0], rd.get("pos2") or [0, 0, 0]
        return p1 == p2

    def parent_chain(rid):
        out, cur = [], rid
        for _ in range(32):
            rd = RG.get(cur)
            pid = rd.get("parentIsland") if rd else None
            if not pid:
                break
            out.append(pid)
            cur = pid
        return out

    PLAYER_RE = re.compile(r"^(개인섬_|길드섬_)")
    fishing_regions = [r for r in regionfish if not PLAYER_RE.match(r)]

    def pool_for(rid, scenario="낮맑음"):
        """FishingListener 의 풀 조립을 복제 (기본 + 부모체인 기본 + 시간/날씨 sublist + env)."""
        names = OrderedDict()
        for n in regionfish.get(rid, {}).get("기본", []):
            names[n] = "기본"
        for pid in parent_chain(rid):
            for n in regionfish.get(pid, {}).get("기본", []):
                names.setdefault(n, f"상속({pid})")
        sub = regionfish.get(rid, {})
        if scenario == "낮맑음":
            for n in sub.get("낮맑음", []):
                names.setdefault(n, "낮맑음")
        else:  # 밤맑음
            for n in envfish.get("밤", []):
                names.setdefault(n, "env:밤")
            for n in sub.get("밤맑음", []):
                names.setdefault(n, "밤맑음")
            for n in sub.get("낮맑음", []):
                names.setdefault(n, "낮맑음(25%누수)")
        return names

    def grade_buckets(names, maxg=9):
        """등급→어종목록. 범위등급('E~S')은 여러 버킷에 들어간다."""
        buckets = defaultdict(list)
        for n in names:
            d = fishdefs.get(n)
            if not d or not d.get("grade"):
                continue
            g = d["grade"]
            if "~" in g:
                a, b = g.split("~")
                lo, hi = GRADE_RANK.get(a, 0), GRADE_RANK.get(b, 0)
            else:
                lo = hi = GRADE_RANK.get(g, 0)
            hi = min(hi, maxg)
            for rank in range(max(lo, 1), hi + 1):
                buckets[list(GRADE_RANK)[rank - 1]].append(n)
        return buckets

    # ============================================================== 00 읽는법
    ws = wb.create_sheet("00_읽는법")
    ws.sheet_properties.tabColor = "1F3864"
    ws["A1"] = "바르칸 열도 — 밸런스 작업용 통합 자료"
    ws["A1"].font = Font(bold=True, size=16)
    intro = [
        "",
        f"생성 시각: {datetime.datetime.now():%Y-%m-%d %H:%M}",
        "",
        "■ 이 파일은 무엇인가",
        "  게임의 라이브 데이터(운영 중인 서버가 실제로 읽는 파일)에서 밸런스 관련 수치를 전부 긁어",
        "  하나로 모은 것이다. 문서를 사람이 옮겨 적은 게 아니라 기계가 뽑았으므로 서버의 현재 상태와",
        "  일치한다. 반대로 말해 이 파일을 고쳐도 게임은 바뀌지 않는다 — 제안은 노란 「제안」 열에 적을 것.",
        "",
        "■ 작업 방법",
        "  1) 각 시트의 회색 열 = 현재값(읽기 전용 취급). 「제안」/「메모」 열에만 적는다.",
        "  2) 수치를 바꾸자는 제안은 '왜'를 한 줄이라도 같이 적어 주면 반영 판단이 빨라진다.",
        "  3) 90_점검표 시트는 기계가 미리 찾아낸 이상 징후 목록이다. 여기서부터 보는 게 빠르다.",
        "",
        "■ 용어",
        "  등급        E < D < C < B < A < S < M < L < G  (9단계, G가 최상)",
        "  캐스트      낚싯대를 한 번 던져 입질까지 간 1회. 확률의 기본 단위.",
        "  피티(PRD)   못 나온 캐스트마다 확률이 선형으로 오르는 보정. 그래서 '기본확률'과",
        "              '실제 출현율'이 크게 다르다(둘 다 표기했다).",
        "  크기점수    (잡은 크기 − 최소) / (최대 − 최소) × 100. 판매가·경험치에 곱해진다.",
        "  스탯 19종   행운·난이도·등급업·크기·크리확률·크리배율·경험치·더블/트리플찬스·판매보너스·",
        "              도망감소·내구보존·등급특화 + 작살 전용(수중호흡·수영속도·공격력·공격속도·돌진쿨감·야간투시)",
        "",
        "■ 권위 소스 (dev 기준 — 코드/데이터가 dev → prod 방향으로 배포되므로 dev가 최신이다)",
        f"  JSON  : {JSON_ROOT}",
        f"  Java  : {JAVA_ROOT}",
        "  코드상수: stats.db catalog_version (요리·통발·작물·드릴·섬/길드가격)",
        "  운영 서버(prod) 기준으로 뽑고 싶으면 prod 의 BlockShip 폴더를 내려받아 --json 으로 지정한다.",
        "",
        "■ 재생성",
        "  python3 .claude/skills/balance-audit/scripts/gen_balance_workbook.py --out <경로.xlsx>",
        "",
    ]
    for i, line in enumerate(intro, 2):
        ws.cell(row=i, column=1, value=line)
    ws.column_dimensions["A"].width = 110

    if CAT_TS:
        r0 = len(intro) + 3
        ws.cell(row=r0, column=1, value="■ 코드상수 스냅샷 시점 (dev 서버 마지막 부팅 기준)").font = Font(bold=True)
        for j, (k, t) in enumerate(sorted(CAT_TS.items()), r0 + 1):
            ws.cell(row=j, column=1, value=f"  {k:<14} {t:%Y-%m-%d %H:%M}")

    # ============================================================== 01 핵심상수
    rows = []
    for g, base, gate in roll_order:
        rows.append(["등급 기본확률", g, base, "%", f"레벨게이트 {gate or '-'}", "피티 전 순수 base"])
    rows.append(["등급 기본확률", "E", "잔여", "%", "-", "위 8종이 모두 실패하면 E"])
    for g, v in gprice.items():
        rows.append(["등급 기본가", g, v, "원", "크기점수 0 → ×0.5 / 100 → ×1.0", "판매 기본가"])
    for g, v in JC["base_exp"].items():
        rows.append(["등급 기본경험치", g, v, "XP", "×(0.5+크기점수/100)", ""])
    for mins, mult in JC["freshness"]:
        rows.append(["신선도 배율", f"~{mins}분", mult, "배", "", "보관시간 초과분"])
    rows.append(["신선도 배율", f">{JC['freshness'][-1][0] if JC['freshness'] else 180}분", JC["freshness_tail"], "배", "", ""])
    for lv, gnum in JC["grade_unlock"]:
        rows.append(["최대등급 해금", f"Lv.{lv}", list(GRADE_RANK)[gnum - 1], "등급", "", "이 레벨부터 해당 등급 롤 개방"])
    formulas = [
        ("판매가 공식", "등급기본가 × (0.5 + 크기점수×0.5/100) × 크리배율 × (1+판매보너스%) × 신선도"),
        ("경험치 공식", "등급기본XP × (0.5 + 크기점수/100) × (1 + (장비경험치% + 환경경험치%)/100) × 크리배율"),
        ("등급업 확률", "장비 등급업 + 환경 등급업 (상한 없음 — 100%에서 자연포화)"),
        ("콤보 보너스", "floor(콤보/5) % — 경험치에만 적용. 더블/트리플엔 절대 얹지 않는다"),
        ("더블/트리플", "독립 롤. 더블 100% 초과분은 ×0.5 로 트리플에 이월"),
        ("크기점수", "floor((크기 − 최소) / (최대 − 최소) × 100), 초대형은 100 초과 가능"),
        ("돈 상한", "1,000,000,000,000,000 (1e15)"),
        ("행운 스탯", "등급 롤 확률 전체에 ×(100+행운)/100"),
    ]
    for k, v in formulas:
        rows.append(["공식", k, v, "", "", ""])
    sheet(
        "01_핵심상수", "핵심 상수·공식 (Java 하드코딩)",
        "수식·확률은 JSON이 아니라 코드에 박혀 있다. 여기 값을 바꾸려면 개발자 작업이 필요하다.",
        ["구분", "항목", "값", "단위", "조건", "메모"], rows,
        {"값": 62, "조건": 34, "메모": 32}, tab="1F3864",
    )

    # ============================================================== 02 레벨곡선
    need = JC["need_table"]
    rows, cum = [], 0
    for i, n in enumerate(need, 1):
        cum += n
        prev = need[i - 2] if i >= 2 else n
        rows.append([i, n, cum, round(n / prev, 3) if prev else "", ""])
    sheet(
        "02_레벨곡선", "낚시 레벨 필요경험치 (Lv.1~100)",
        "누적 = 그 레벨을 달성하기까지 필요한 총 XP. 배수 = 전 레벨 대비 벽. 베타 종점은 Lv.70 설계.",
        ["레벨", "필요XP", "누적XP", "전레벨대비", "제안"], rows,
        {"레벨": 8, "필요XP": 12, "누적XP": 14, "전레벨대비": 12, "제안": 30}, tab="1F3864",
        numfmt={"필요XP": "#,##0", "누적XP": "#,##0"},
    )

    rows = [[lv, k, v, ""] for lv, k, v in parse_level_bonus(JC["level_bonus_src"])]
    sheet(
        "03_레벨보너스", "레벨 마일스톤 스탯 (levelBonus)",
        "장비와 무관하게 레벨만으로 붙는 보너스. 5레벨 단위 계단.",
        ["레벨", "항목", "증가", "제안"], rows, {"제안": 30}, tab="1F3864",
    )

    # ============================================================== 04 지역
    rows = []
    for rid in sorted(set(list(RG.keys()) + fishing_regions)):
        if PLAYER_RE.match(rid):
            continue
        rd = RG.get(rid, {})
        pool = pool_for(rid) if rid in regionfish else {}
        own = regionfish.get(rid, {})
        rows.append([
            rid,
            rd.get("displayName", ""),
            rd.get("parentIsland", ""),
            rd.get("world", ""),
            rd.get("requiredLevel", ""),
            "등록안됨" if rid not in RG else ("스텁(좌표없음)" if is_stub(rid) else "정상"),
            len(own.get("기본", [])),
            len(pool),
            len(own.get("통발", [])),
            len(own.get("밤맑음", [])) + len(own.get("밤비", [])),
            len(drops.get(rid, [])),
            "예" if rid in drops else ("상속" if any(p in drops for p in parent_chain(rid)) else "없음"),
            ",".join(rd.get("allowedWeathers", []) or []),
            ",".join(rd.get("excludedWeathers", []) or []),
            "",
        ])
    sheet(
        "04_지역목록", "지역 마스터 (낚시터 + 상위 지역)",
        "★상태=스텁/등록안됨 인 지역은 데이터엔 있지만 게임에 실제로 갈 수 없다 — 그 지역 수치로 결론내지 말 것.",
        ["지역id", "표시명", "부모지역", "월드", "요구레벨", "상태", "기본어종수", "실풀크기(상속포함)",
         "통발어종", "밤전용어종", "재료종수", "드롭테이블", "허용날씨", "제외날씨", "메모"],
        rows, {"지역id": 22, "표시명": 14, "부모지역": 14, "메모": 26}, tab="C00000",
    )

    # ============================================================== 05 지역별 재료확률 ★
    all_mats = []
    for tbl in drops.values():
        for d in tbl:
            if d["matId"] not in all_mats:
                all_mats.append(d["matId"])
    all_mats.sort(key=lambda m: -sum(d["chance"] for t in drops.values() for d in t if d["matId"] == m))

    rows = []
    for rid in sorted(drops.keys()):
        tbl = {d["matId"]: d["chance"] for d in drops[rid]}
        total = sum(tbl.values())
        row = [rid, "스텁/미등록" if is_stub(rid) else "정상", len(tbl), round(total, 1),
               round(total / 100.0, 3)]
        for m in all_mats:
            row.append(tbl.get(m, None))
        rows.append(row + [""])
    # 상속으로만 재료가 나오는 낚시 지역도 표시
    for rid in sorted(fishing_regions):
        if rid in drops:
            continue
        src = next((p for p in parent_chain(rid) if p in drops), None)
        tbl = {d["matId"]: d["chance"] for d in drops.get(src, [])} if src else {}
        total = sum(tbl.values())
        rows.append(
            [rid, f"상속←{src}" if src else "재료 안 나옴", len(tbl), round(total, 1), round(total / 100.0, 3)]
            + [tbl.get(m, None) for m in all_mats] + [""]
        )
    ws = sheet(
        "05_지역별재료확률", "지역별 재료 드롭 확률 (%/캐스트, 독립 롤)",
        "각 재료가 서로 독립으로 굴러간다(합이 100%일 필요 없음). 값=물고기 1마리 잡을 때 그 재료가 나올 %. "
        "자기 테이블이 없으면 부모 지역 테이블을 상속하며, 병합이 아니라 대체다.",
        ["지역", "지역상태", "재료종수", "확률합", "기대개수/캐스트"] + [mats.get(m, {}).get("name", m) for m in all_mats] + ["메모"],
        rows, {"지역": 20, "지역상태": 14, "메모": 24}, tab="C00000",
    )
    for r in range(5, 5 + len(rows)):
        for c in range(6, 6 + len(all_mats)):
            ws.cell(row=r, column=c).number_format = '0.#"%"'

    # ============================================================== 06 재료 마스터
    mat_use = defaultdict(list)
    mat_qty = Counter()
    for rid, rc in recipes.items():
        for i in rc.get("ingredients", []):
            if i.get("kind") == "custom":
                mat_use[i["typeOrMatId"]].append(rc.get("displayName", rid))
                mat_qty[i["typeOrMatId"]] += i.get("qty", 1)
    mat_src = defaultdict(list)
    for rid, tbl in drops.items():
        for d in tbl:
            mat_src[d["matId"]].append(f"{rid} {d['chance']}%")
    # recipes.json 결과물의 mat: 로어 마커 → 조합으로 얻는 재료
    mat_from_recipe = {}
    for rid, rc in recipes.items():
        for line in (rc.get("result") or {}).get("lore", []) or []:
            mm = re.search(r"mat:(\S+)", line)
            if mm:
                mat_from_recipe[mm.group(1)] = rid
    # 드릴 채굴 산출
    drill_min = {
        o.get("mineral") for o in ((CAT.get("drill") or {}).get("ores") or {}).values() if o.get("mineral")
    }
    print("Java 소스 스캔(재료 획득경로)...")
    where = java_mention_index()

    mat_paths = {}
    rows = []
    for mid, m in mats.items():
        srcs = mat_src.get(mid, [])
        paths = []
        if srcs:
            paths.append("낚시 드롭")
        if mid in mat_from_recipe:
            paths.append(f"조합({mat_from_recipe[mid]})")
        if mid in drill_min or m.get("name") in drill_min:
            paths.append("드릴 채굴")
        _hits, labels = where(mid)
        for lab in labels:
            if lab not in paths and lab not in ("상점", "요리 재료"):
                paths.append(lab)
        mat_paths[mid] = paths
        rows.append([
            mid, m.get("name", ""), m.get("mcItem", ""),
            len(srcs), ", ".join(srcs) if srcs else "",
            " / ".join(paths) if paths else "★없음(획득 불가 후보)",
            len(mat_use.get(mid, [])), mat_qty.get(mid, 0),
            ", ".join(sorted(set(mat_use.get(mid, [])))[:8]),
            m.get("desc", ""), "",
        ])
    rows.sort(key=lambda r: (-r[3], -r[6]))
    sheet(
        "06_재료마스터", f"조합 재료 마스터 ({len(mats)}종)",
        "획득경로는 JSON(드롭테이블·레시피) + 코드(RecipeLoader 주입 레시피·드릴·보스 등) 양쪽을 스캔한 결과다. "
        "'없음'이면 진짜 획득 불가 후보. 사용처 0 = 쓸 데 없는 재료.",
        ["재료id", "이름", "MC아이템", "획득지역수", "낚시 획득처(지역 확률)", "획득경로(전체)",
         "사용레시피수", "총소요량", "사용처(일부)", "설명", "메모"],
        rows, {"낚시 획득처(지역 확률)": 40, "획득경로(전체)": 26, "사용처(일부)": 34, "설명": 28, "메모": 22},
        tab="C00000",
    )

    # ============================================================== 07 어종 등급확률
    print("등급 PRD 시뮬레이션...")
    scen_rows = []
    for scen in ("낮맑음", "밤맑음"):
        for rid in sorted(fishing_regions):
            names = pool_for(rid, scen)
            bk = grade_buckets(names)
            rates = sim_grade_rates(set(bk.keys()), sim_casts, roll_order=roll_order)
            row = [scen, rid, "스텁/미등록" if is_stub(rid) else "정상", len(names)]
            for g in "EDCBASMLG":
                row.append(len(bk.get(g, [])))
            for g in "EDCBASMLG":
                row.append(round(rates.get(g, 0), 4) if bk.get(g) else 0)
            exp_price = sum(
                rates.get(g, 0) / 100.0 * gprice.get(g, 0) for g in "EDCBASMLG" if bk.get(g)
            )
            exp_xp = sum(
                rates.get(g, 0) / 100.0 * JC["base_exp"].get(g, 0) for g in "EDCBASMLG" if bk.get(g)
            )
            row += [round(exp_price), round(exp_xp, 2), ""]
            scen_rows.append(row)
    sheet(
        "07_지역별등급확률", "지역별 등급 실제 출현율 (PRD 시뮬)",
        f"몬테카를로 {sim_casts:,}캐스트, 스탯 0·Lv.100(전등급 개방)·B/C 해금 완료 기준. "
        "★그 지역 풀에 없는 등급은 롤을 건너뛰므로 피티가 보존되고, 남은 등급의 실확률이 올라간다. "
        "기대판매가는 크기점수 100(=×1.0) 기준 상한값이다.",
        ["시나리오", "지역", "상태", "풀크기"]
        + [f"{g}종수" for g in "EDCBASMLG"]
        + [f"{g}실확률%" for g in "EDCBASMLG"]
        + ["기대판매가(최대)", "기대XP", "메모"],
        scen_rows, {"지역": 20, "시나리오": 10, "메모": 20}, tab="C00000",
        numfmt={f"{g}실확률%": "0.0000" for g in "EDCBASMLG"} | {"기대판매가(최대)": "#,##0"},
    )

    # ============================================================== 08 지역별 어종 확률
    rows = []
    for rid in sorted(fishing_regions):
        names = pool_for(rid, "낮맑음")
        bk = grade_buckets(names)
        rates = sim_grade_rates(set(bk.keys()), sim_casts, roll_order=roll_order)
        for n, origin in names.items():
            d = fishdefs.get(n)
            if not d:
                continue
            g = d.get("grade", "")
            gl = [x for x in "EDCBASMLG" if n in bk.get(x, [])]
            p = sum(rates.get(x, 0) / len(bk[x]) for x in gl if bk.get(x))
            rows.append([
                rid, n, g, origin,
                d.get("minSize"), d.get("maxSize"), d.get("time", ""), d.get("weather", ""),
                ",".join(gl), round(p, 5),
                round(1 / (p / 100), 1) if p > 0 else "",
                gprice.get(gl[-1], "") if gl else "",
                "",
            ])
    rows.sort(key=lambda r: (r[0], -(r[9] or 0)))
    sheet(
        "08_지역별어종확률", "지역별 어종 출현확률 (낮·맑음 기준)",
        "확률 = 그 등급이 나올 실확률 ÷ 그 등급 풀의 어종 수(균등 추첨). 범위등급 어종은 여러 등급에 걸쳐 있어 합산했다. "
        "출처=기본/상속(부모지역)/시간·날씨 서브리스트. 밤 전용 어종은 낮 풀에 아예 없다.",
        ["지역", "어종", "등급표기", "풀 출처", "최소cm", "최대cm", "출현시간", "출현날씨",
         "실제등급버킷", "확률%", "평균 몇캐스트당 1마리", "등급기본가", "메모"],
        rows, {"지역": 18, "어종": 18, "풀 출처": 16, "메모": 18}, tab="C00000",
        numfmt={"확률%": "0.00000", "평균 몇캐스트당 1마리": "#,##0.0", "등급기본가": "#,##0"},
    )

    # ============================================================== 09 어종 마스터
    fish_regions = defaultdict(list)
    for rid, subs in regionfish.items():
        if PLAYER_RE.match(rid):
            continue
        for sub, lst in subs.items():
            for n in lst:
                fish_regions[n].append(f"{rid}/{sub}")
    for cond, lst in envfish.items():
        for n in lst:
            fish_regions[n].append(f"env:{cond}")
    rows = []
    for n, d in fishdefs.items():
        g = d.get("grade", "")
        top = g.split("~")[-1] if "~" in g else g
        rows.append([
            n, g, d.get("minSize"), d.get("maxSize"),
            round((d.get("maxSize") or 0) - (d.get("minSize") or 0), 1),
            d.get("time", ""), d.get("weather", ""), d.get("quest", ""),
            gprice.get(top, ""), round(gprice.get(top, 0) * 0.5) if top in gprice else "",
            JC["base_exp"].get(top, ""),
            len(fish_regions.get(n, [])), ", ".join(fish_regions.get(n, [])[:6]), "",
        ])
    rows.sort(key=lambda r: (GRADE_RANK.get(str(r[1]).split("~")[-1], 0), r[0]))
    sheet(
        "09_어종마스터", f"어종 마스터 ({len(fishdefs)}종)",
        "판매가는 크기점수 100(최대) / 0(최소) 기준. 배치지역 0 = 어디에도 안 배치된 어종.",
        ["어종", "등급표기", "최소cm", "최대cm", "폭", "출현시간", "출현날씨", "퀘스트게이트",
         "최대판매가", "최소판매가", "기본XP", "배치지역수", "배치지역(일부)", "메모"],
        rows, {"어종": 20, "배치지역(일부)": 40, "메모": 18}, tab="C00000",
        numfmt={"최대판매가": "#,##0", "최소판매가": "#,##0"},
    )

    # ============================================================== 10~12 장비
    def equip_sheet(sname, title, types, note, tab):
        rows = []
        for t in types:
            for name, spec in parts.get(t, {}).items():
                pd = parse_part(spec)
                rc = rec_by_rod.get(name) if t == "낚싯대" else rec_by_part.get((t, name))
                if rc is None:
                    rc = rec_by_part.get((t, name)) or rec_by_rod.get(name)
                st = pd["스탯"]
                stat_sum = sum(v for v in st.values() if isinstance(v, (int, float)))
                unlock = round(pd["가격"] * 0.5) if rc and rc.get("locked") else ""
                row = [
                    t, pd["이름"], pd["등급"], GRADE_RANK.get(pd["등급"], 0),
                    pd["레벨제한"], pd["출처"], pd["가격"], pd["내구"],
                    stat_sum, round(pd["가격"] / stat_sum) if stat_sum else "",
                ]
                row += [st.get(k) for k in STAT_ORDER]
                row += [
                    rc.get("id") if rc else "레시피없음",
                    "예" if rc and rc.get("locked") else ("아니오" if rc else ""),
                    unlock,
                    (rc or {}).get("village", ""),
                    len((rc or {}).get("ingredients", [])),
                    ing_text(rc),
                ]
                row += ing_cols(rc)
                row.append("")
                rows.append(row)
        rows.sort(key=lambda r: (r[0], r[4], r[3], r[6]))
        headers = (
            ["부위", "이름", "등급", "등급순위", "레벨제한", "출처(상점)", "상점가", "내구도",
             "스탯합", "가격/스탯"]
            + STAT_ORDER
            + ["레시피id", "잠김", "해금비(가격×0.5)", "제작마을", "재료종수", "재료(전체)"]
            + sum([[f"재료{i}", f"수량{i}"] for i in range(1, 7)], [])
            + ["메모"]
        )
        return sheet(
            sname, title, note, headers, rows,
            {"이름": 22, "출처(상점)": 14, "재료(전체)": 52, "메모": 20},
            tab=tab,
            numfmt={"상점가": "#,##0", "해금비(가격×0.5)": "#,##0", "가격/스탯": "#,##0"},
        )

    NOTE_EQ = (
        "★상점가 ≠ 실제 획득비용. 카테고리 최저사양 1개만 돈으로 사고, 나머지는 상점에서 '레시피 해금비'(보통 "
        "가격×0.5, 1회성)만 내고 그 다음부터는 재료로 제작한다. 진짜 반복비용은 재료다(재료는 낚시 드롭 + 광질). "
        "빈 스탯칸 = 그 스탯 없음. 스탯합/가격당스탯은 이질적 스탯을 단순 합친 거친 지표이니 같은 부위·같은 등급 안에서만 비교할 것."
    )
    equip_sheet("10_낚싯대", f"낚싯대 ({len(parts.get('낚싯대', {}))}종)", ["낚싯대"], NOTE_EQ, "375623")
    equip_sheet("11_작살", f"작살 ({len(parts.get('작살', {}))}종)",
                ["작살"], NOTE_EQ + " 작살은 수중 근접 사냥용이라 전용 스탯(수중호흡·수영속도·공격력·공격속도·돌진쿨감·야간투시)을 쓴다.", "375623")
    equip_sheet("12_부품", f"부품 ({sum(len(parts.get(t, {})) for t in ['릴', '줄', '바늘', '미끼', '찌'])}종)",
                ["릴", "줄", "바늘", "미끼", "찌"],
                NOTE_EQ + " 부품은 내구도가 있고 0이 되면 고장(스탯 미적용, 수리 필요)이다. 낚싯대는 내구 소모 없음.", "375623")

    # ============================================================== 13 레시피 전체
    rows = []
    for rid, rc in recipes.items():
        ings = rc.get("ingredients", [])
        row = [
            rid, rc.get("category", ""), rc.get("displayName", ""),
            rc.get("resultMode", ""), rc.get("resultPartType", "") or rc.get("rodPartName", ""),
            "예" if rc.get("locked") else "아니오",
            rc.get("village", ""), rc.get("drillTier", 0),
            len(ings), sum(i.get("qty", 1) for i in ings),
            ing_text(rc),
        ]
        row += ing_cols(rc, 8)
        row.append("")
        rows.append(row)
    rows.sort(key=lambda r: (r[1], r[0]))
    sheet(
        "13_레시피전체", f"조합 레시피 전체 ({len(recipes)}건)",
        "kind=custom → 커스텀 재료(mat:) / item → 바닐라 아이템 / fish → 해당 등급 이상 물고기 / herbany → 허브류 아무거나. "
        "잠김=예 는 상점에서 해금비를 내야 제작창에 뜬다.",
        ["레시피id", "카테고리", "결과물", "결과형태", "결과부위", "잠김", "판매마을", "드릴티어",
         "재료종수", "재료총개수", "재료(전체)"]
        + sum([[f"재료{i}", f"수량{i}"] for i in range(1, 9)], [])
        + ["메모"],
        rows, {"결과물": 22, "재료(전체)": 60, "메모": 18}, tab="375623",
    )

    # ============================================================== 14 강화
    enh = CAT.get("enhance", {})
    cost, succ, down = enh.get("cost", []), enh.get("success", []), enh.get("down", [])
    pearl, sl, sc = enh.get("pearl", []), enh.get("sacrificeLevel", []), enh.get("sacrificeCount", [])
    ckpt = set(enh.get("checkpoint", []))
    rows = []
    for lv in range(1, max(len(cost), len(succ))):
        s = succ[lv] if lv < len(succ) else None
        rows.append([
            f"+{lv - 1} → +{lv}", lv,
            cost[lv] if lv < len(cost) else None,
            s, round(100 / s, 2) if s else "",
            round((cost[lv] if lv < len(cost) else 0) * 100 / s) if s else "",
            down[lv] if lv < len(down) else None,
            pearl[lv] if lv < len(pearl) else None,
            "예" if (lv - 1) in ckpt else "",
            (sl[lv] if lv < len(sl) else 0) or "",
            (sc[lv] if lv < len(sc) else 0) or "",
            "",
        ])
    sheet(
        "14_강화비용", "낚싯대 강화 단계표",
        "실패하면 하락% 확률로 한 단계 내려간다(체크포인트=예 인 단계 아래로는 안 내려감). "
        "기대비용은 '단순 1/성공률 × 비용'이라 하락 손실을 포함하지 않은 하한값이다. 제물=같은 낚싯대(해당 강화도) 소모.",
        ["구간", "목표+", "골드비용", "성공률%", "기대시도횟수", "기대골드(하한)", "실패시하락%",
         "별빛진주", "체크포인트", "제물 강화도", "제물 개수", "메모"],
        rows, {"구간": 12, "메모": 24}, tab="375623",
        numfmt={"골드비용": "#,##0", "기대골드(하한)": "#,##0"},
    )

    # 강화 스탯 증가표
    rows = []
    for rod in EN.get("order", []):
        t = EN.get("table", {}).get(rod)
        if not t:
            continue
        cum = defaultdict(float)
        for lv in range(1, int(t.get("max", 0)) + 1):
            s = t.get("levels", {}).get(str(lv), "")
            step = {}
            for tok in s.split(","):
                if ":" in tok:
                    k, v = tok.split(":", 1)
                    step[k.strip()] = float(v)
                    cum[k.strip()] += float(v)
            rows.append(
                [rod, t.get("max"), lv, s]
                + [step.get(k) for k in STAT_ORDER]
                + [round(cum[k], 1) if cum[k] else None for k in STAT_ORDER]
                + [""]
            )
    sheet(
        "15_강화스탯", "낚싯대별 강화 단계 스탯 증가",
        "왼쪽=그 단계에서 붙는 증가분, 오른쪽(누적)=그 단계까지 총합. max=그 낚싯대의 강화 상한.",
        ["낚싯대", "최대강화", "단계", "원문"]
        + [f"+{k}" for k in STAT_ORDER]
        + [f"누적{k}" for k in STAT_ORDER]
        + ["메모"],
        rows, {"낚싯대": 22, "원문": 30, "메모": 16}, tab="375623",
    )

    # ============================================================== 16 환경보너스
    rows = []
    keys = sorted({k for v in list(EB.get("weathers", {}).values()) + list(EB.get("times", {}).values()) for k in v})
    for kind, table in (("날씨", EB.get("weathers", {})), ("시간대", EB.get("times", {}))):
        for name, st in table.items():
            rows.append([kind, name] + [st.get(k) for k in keys] + [sum(st.values()), ""])
    sheet(
        "16_환경보너스", "날씨·시간대 보너스",
        "장비 스탯과 합연산으로 더해진다. 시간대 구간: 새벽 22000~999 / 낮 1000~10999 / 저녁 11000~14999 / 밤 15000~21999.",
        ["구분", "이름"] + keys + ["합계", "메모"], rows, {"메모": 24}, tab="C55A11",
    )

    # ============================================================== 17 요리
    dishes = CAT.get("dishes", {})
    rows = []
    for did, d in sorted(dishes.items(), key=lambda kv: (kv[1].get("purpose", ""), kv[1].get("tier", 0), kv[0])):
        ings = d.get("ingredients", [])
        row = [
            did, re.sub(r"§.", "", d.get("name", "")), d.get("purpose", ""), d.get("tier", ""),
            d.get("base", ""), d.get("cookTimeSec"), d.get("durationSec"),
            d.get("exp"), d.get("size"), d.get("gradeup"), d.get("crit"), d.get("dbl"),
            d.get("sellBonus"), d.get("escape"), d.get("difficulty"), d.get("heal"),
            round(sum(float(d.get(k) or 0) for k in ("exp", "size", "gradeup", "crit", "dbl", "sellBonus", "escape")), 1),
            d.get("submitPoints"), d.get("sellPrice"),
            len(ings),
            " + ".join(f"{i.get('displayName')} x{i.get('qty')}" for i in ings),
        ]
        for i in range(4):
            if i < len(ings):
                row += [ings[i].get("displayName"), ings[i].get("qty"), ings[i].get("kind")]
            else:
                row += ["", None, ""]
        row.append("")
        rows.append(row)
    sheet(
        "17_요리", f"요리 ({len(dishes)}종)",
        "목적: buff=먹는 버프 전용(제출 0점) / submit=제출 전용(버프 0) / heal=즉시회복 / sell=판매용. "
        "지속=버프 지속초. 버프는 1종만 활성. 완벽 요리(perfect)면 효과 +50~100%.",
        ["요리id", "표시명", "목적", "티어", "베이스아이템", "조리초", "지속초",
         "경험치", "크기", "등급업", "크리확률", "더블", "판매보너스", "도망감소", "난이도", "회복",
         "버프합", "제출점수", "판매가", "재료종수", "재료(전체)"]
        + sum([[f"재료{i}", f"수량{i}", f"종류{i}"] for i in range(1, 5)], [])
        + ["메모"],
        rows, {"요리id": 16, "표시명": 16, "재료(전체)": 50, "메모": 18}, tab="C55A11",
        numfmt={"제출점수": "#,##0", "판매가": "#,##0"},
    )

    # ============================================================== 18 통발
    traps = CAT.get("traps", {})
    rows = []
    for key, t in sorted(traps.items()):
        ings = t.get("ingredients", [])
        rid = t.get("region", "")
        trapfish = regionfish.get(rid, {}).get("통발", [])
        rows.append([
            rid, t.get("regionLabel", ""), t.get("variant", ""),
            re.sub(r"&.", "", t.get("trapName", "")),
            "스텁/미등록" if is_stub(rid) else "정상",
            t.get("maxDur"), t.get("waitSec"), round((t.get("waitSec") or 0) / 60, 1),
            t.get("luck"), t.get("unlockPrice"), t.get("recipeId"),
            len(ings), " + ".join(f"{i.get('krName')} x{i.get('qty')}" for i in ings),
            len(trapfish), ", ".join(trapfish), "",
        ])
    sheet(
        "18_통발", f"통발 ({len(traps)}종)",
        "지역마다 표준/튼튼/속성/행운 변종. 내구=사용 가능 회수, 대기=1회 수확까지 초. "
        "★지역상태가 스텁/미등록이면 레시피는 팔리는데 설치할 곳이 없다.",
        ["지역", "지역표시명", "변종", "통발명", "지역상태", "내구(회수)", "대기초", "대기분",
         "행운%", "해금가", "레시피id", "재료종수", "재료", "전용어종수", "전용어종", "메모"],
        rows, {"통발명": 20, "재료": 44, "전용어종": 28, "메모": 16}, tab="C55A11",
        numfmt={"해금가": "#,##0"},
    )

    # ============================================================== 19 작물/채집/드릴
    crops = CAT.get("crops", {})
    rows = [[
        c.get("id"), re.sub(r"§.", "", c.get("name", "")), c.get("growSec"),
        round((c.get("growSec") or 0) / 60, 1), c.get("outQty"), c.get("matId"),
        c.get("outMat"), len(c.get("stageVariants", [])), c.get("furnitureId"), "",
    ] for c in crops.values()]
    sheet(
        "19_특수작물", f"특수 작물 ({len(crops)}종)",
        "요리 재료용. 성장시간은 고정(날씨/비료 영향 없음). 섬 레벨별 심을 수 있는 개수 상한은 21_섬길드가격 시트 참조.",
        ["작물id", "표시명", "성장초", "성장분", "수확개수", "산출 재료id", "MC아이템", "성장단계수", "가구id", "메모"],
        rows, {"메모": 26}, tab="C55A11",
    )

    forage = CAT.get("forage", {}) or read_json("forage-types.json", {})
    rows = [[
        k, v.get("name", ""), v.get("region", ""), v.get("rarity", ""),
        v.get("cooldownSec"), round((v.get("cooldownSec") or 0) / 3600, 2), "",
    ] for k, v in sorted(forage.items(), key=lambda kv: (kv[1].get("rarity", ""), kv[1].get("region", "")))]
    sheet(
        "20_채집", f"채집물 ({len(forage)}종)",
        "쿨타임은 유저별(노드는 남아 있고 그 유저에게만 안 보임). 지역은 배치 테마 이름이며 regions.json 지역id와 다르다.",
        ["노드id", "이름", "배치테마", "희귀도", "쿨타임초", "쿨타임시간", "메모"],
        rows, {"노드id": 30, "메모": 26}, tab="C55A11",
    )

    drill = CAT.get("drill", {})
    rows = []
    for t in (drill.get("tiers") or {}).values():
        rows.append(["드릴", t.get("tier"), re.sub(r"§.", "", t.get("name", "")), t.get("speed"),
                     "", "", "", "", re.sub(r"§.", "", t.get("desc", "")), ""])
    for blk, o in (drill.get("ores") or {}).items():
        rows.append(["광맥", o.get("requiredTier"), o.get("label", ""), "", blk,
                     o.get("mineral"), o.get("breakTicks"), o.get("regenSeconds"),
                     f"{o.get('lootMin')}~{o.get('lootMax')}개", ""])
    sheet(
        "21_드릴광질", "드릴 · 광맥",
        "T1 흑정석/T2 자수정. 채굴 파괴시간(틱)은 드릴 speed로 나뉘고, regen 후 광맥이 되살아난다. "
        "★regions.json '광산' 지역은 채굴처가 아니다(실제 채굴은 mine 월드와 레드_로드).",
        ["구분", "요구티어", "이름/라벨", "속도배율", "블록", "산출광물", "파괴틱", "재생초", "산출량", "메모"],
        rows, {"메모": 26}, tab="C55A11",
    )

    # ============================================================== 22 퀘스트 보상
    qmain = QJ.get("퀘스트", {})
    rows = []
    for qid, q in qmain.items():
        rows.append([
            qid, re.sub(r"&.", "", q.get("이름", "")), q.get("카테고리", ""), q.get("타입", ""),
            q.get("마을", ""), q.get("난이도"), q.get("필요레벨"),
            q.get("보상돈"), q.get("보상경험치"),
            " / ".join(q.get("목표", [])),
            q.get("선행퀘스트", ""), q.get("다음퀘스트", ""),
            q.get("보상칭호", ""), ", ".join(q.get("보상재료", []) or []) if isinstance(q.get("보상재료"), list) else q.get("보상재료", ""),
            "",
        ])
    rows.sort(key=lambda r: (r[2], r[6] or 0, r[0]))
    sheet(
        "22_퀘스트보상", f"퀘스트 보상 ({len(qmain)}건)",
        "설계 의도는 '경험치는 퀘스트에서, 낚시는 적게'. 목표 문법: fish|어종|등급|수량, fish_fresh|…|신선도|크기, "
        "material|재료|수량, sell|금액, area|지역, npc|이름 등.",
        ["퀘스트id", "이름", "카테고리", "타입", "마을", "난이도", "필요레벨", "보상돈", "보상XP",
         "목표", "선행", "다음", "보상칭호", "보상재료", "메모"],
        rows, {"이름": 26, "목표": 34, "메모": 18}, tab="7030A0",
        numfmt={"보상돈": "#,##0", "보상XP": "#,##0"},
    )

    dl = QJ.get("일일", {})
    lv = QJ.get("난이도레벨", {})
    rows = []
    for diff, ids in dl.items():
        for qid in ids:
            q = qmain.get(qid, {})
            rows.append(["일일", diff, lv.get(diff, ""), qid, re.sub(r"&.", "", q.get("이름", "")),
                         q.get("보상돈"), q.get("보상경험치"), " / ".join(q.get("목표", [])), ""])
    for qid in QJ.get("주간", []):
        q = qmain.get(qid, {})
        rows.append(["주간", "", "", qid, re.sub(r"&.", "", q.get("이름", "")),
                     q.get("보상돈"), q.get("보상경험치"), " / ".join(q.get("목표", [])), ""])
    sheet(
        "23_일일주간풀", "일일·주간 퀘스트 풀",
        "일일은 난이도 4구간(쉬움/보통/어려움/전문)이며 해금레벨이 다르다. 주간은 레벨 게이팅 방식.",
        ["종류", "난이도", "해금레벨", "퀘스트id", "이름", "보상돈", "보상XP", "목표", "메모"],
        rows, {"이름": 26, "목표": 34, "메모": 18}, tab="7030A0",
        numfmt={"보상돈": "#,##0", "보상XP": "#,##0"},
    )

    # ============================================================== 24 섬/길드 가격
    rows = []
    for scope, tbl in (("개인섬", CAT.get("island_prices", {})), ("길드섬", CAT.get("guild_prices", {}))):
        for k, v in sorted(tbl.items()):
            if isinstance(v, list):
                for i, x in enumerate(v):
                    rows.append([scope, k, i, x, ""])
            else:
                rows.append([scope, k, "", v, ""])
    sheet(
        "24_섬길드가격", "섬 · 길드 업그레이드 가격 사다리",
        "인덱스=업그레이드 레벨. …Limit=그 레벨의 상한 개수, …Price=그 레벨로 올리는 비용. 길드 가격은 개인섬 ×5 원칙.",
        ["구분", "항목", "레벨", "값", "메모"], rows, {"메모": 30}, tab="7030A0",
        numfmt={"값": "#,##0"},
    )

    # ============================================================== 25 제출값
    rows = []
    for k, v in (SV.get("items") or {}).items():
        rows.append(["아이템", k, v, ""])
    for k, v in (SV.get("fishByGrade") or {}).items():
        rows.append(["물고기(등급)", k, v, ""])
    for k, v in (SV.get("rewards") or {}).items():
        rows.append(["순위보상(코인)", k, ", ".join(map(str, v)), ""])
    for d in dishes.values():
        if d.get("submitPoints"):
            rows.append(["요리", re.sub(r"§.", "", d.get("name", "")), d.get("submitPoints"), ""])
    gq = CAT.get("guild_quests", {})
    if gq.get("mult"):
        for i, m in enumerate(gq["mult"]):
            rows.append(["길드임무 목표배수", f"길드Lv.{i}", m, ""])
    sheet(
        "25_제출랭킹", "섬·길드 제출 점수표",
        "제출 = 섬/길드 랭킹 점수. 보상 상한 1,000,000 clamp. 요리 제출점수가 원자재 대비 압도적으로 크므로 여기서 인플레가 난다.",
        ["구분", "항목", "점수/값", "메모"], rows, {"항목": 24, "메모": 30}, tab="7030A0",
        numfmt={"점수/값": "#,##0"},
    )

    # ============================================================== 26 특성 트리
    st = parse_skill_tree()
    rows = [[
        n["숙련"], n["계열"], n["노드id"], n["노드명"], n["종류"], n["최대랭크"],
        n["선행"], n["계열요구"], n["스탯키"], n["랭크당"], n["단위"], n["최대효과"],
        n["효과설명"], "",
    ] for n in st]
    sheet(
        "26_특성트리", f"숙련 특성 트리 ({len(st)}노드)",
        "숙련 레벨 1당 포인트 1. 스탯 노드는 랭크당 선형, 발동(proc) 노드는 PRD 확률 기반 연출형. "
        "계열요구=그 계열에 이미 투자한 포인트 요구량.",
        ["숙련", "계열", "노드id", "노드명", "종류", "최대랭크", "선행노드", "계열요구",
         "스탯키", "랭크당", "단위", "최대효과", "효과설명", "메모"],
        rows, {"효과설명": 44, "메모": 18}, tab="7030A0",
    )

    # ============================================================== 27 상점 물가
    rows = []
    for c in SH.get("categories", []):
        if isinstance(c, dict):
            n_items = len(c.get("items", []) or [])
            rows.append([
                c.get("key", ""), "동적" if c.get("dynamic") else "고정",
                c.get("buy"), c.get("sell"), n_items,
                round(c.get("sell") / c.get("buy"), 3) if c.get("buy") and c.get("sell") else "", "",
            ])
            for it in (c.get("items") or []):
                if isinstance(it, dict):
                    rows.append([
                        f"  └ {c.get('key', '')}", it.get("mcItem") or it.get("item", ""),
                        it.get("buy"), it.get("sell"), "",
                        round(it.get("sell") / it.get("buy"), 3) if it.get("buy") and it.get("sell") else "", "",
                    ])
    sheet(
        "27_상점물가", "일반 상점 매입/매도가",
        "동적 카테고리는 카테고리 단일 가격을 모든 해당 블록에 적용한다. 회수율(=매도/매입)이 1에 가까우면 무한 차익 위험.",
        ["카테고리/아이템", "방식", "매입가(플레이어가 사는 값)", "매도가(파는 값)", "품목수", "회수율", "메모"],
        rows, {"카테고리/아이템": 28, "메모": 26}, tab="7030A0",
        numfmt={"매입가(플레이어가 사는 값)": "#,##0", "매도가(파는 값)": "#,##0"},
    )

    # ============================================================== 28 미니게임 난이도/도주
    mg = read_java("fishing/MinigameTables.java")
    params = {}
    for g, vals in re.findall(r'case "(\w)"\s*-> new Params\(([^)]*)\)', mg):
        nums = [int(x) for x in re.findall(r"-?\d+", vals.replace("gPattern", "0"))]
        params[g] = nums
    fdiff = {}
    m = re.search(r"int fishDifficulty\(String grade\) \{(.*?)\};", mg, re.S)
    if m:
        for g, v in re.findall(r'case "(\w)" -> (\d+)', m.group(1)):
            fdiff[g] = int(v)
    if not params or not fdiff:
        warn("MinigameTables 파싱 실패 — 난이도/도주 시트 부실")

    def derive(grade, size, rod_bonus, esc_red, env_diff=0, env_esc=0):
        p = params.get(grade)
        if not p:
            return None
        size_d = 0 if size < 50 else min(int((size - 50) // 50) + 1, 7)
        net = rod_bonus - fdiff.get(grade, 0) - size_d
        bar = max(12, min(30, 14 - net))
        # Java: 8 + (int) Math.floor(net / 2.0) — 파이썬 // 도 음수에서 같은 방향으로 내림한다
        zone = 8 + (net // 2)
        overflow = 0
        if zone < 1:
            overflow = 1 - zone
            zone = 1
        zone = min(zone, 10)
        bar = max(bar, zone + 2)
        if env_diff > 0:
            zone = max(1, zone - env_diff)
        esc = p[2] - esc_red // 2 - net // 4 + env_esc
        esc = max(1, min(100, esc))
        return net, bar, zone, overflow, esc, p[3]

    rows = []
    for g in "EDCBASMLG":
        p = params.get(g)
        if not p:
            continue
        rows.append([
            "등급 파라미터", g, fdiff.get(g), p[0], p[1], p[2], p[3], p[4], p[5], p[6], p[7], p[8], "",
        ])
    for size, label in ((30, "30cm"), (80, "80cm"), (180, "180cm"), (400, "400cm")):
        d = 0 if size < 50 else min(int((size - 50) // 50) + 1, 7)
        rows.append(["크기 난이도", label, d, "", "", "", "", "", "", "", "", "", "존폭에서 추가 차감"])
    sheet(
        "28_미니게임난이도", "미니게임 파라미터 · 등급별 난이도",
        "존폭=성공 판정 구간(넓을수록 쉬움), 도주기본=시작 도주 확률, 도주증가=미스마다 +%. "
        "장비 '난이도' 스탯이 등급난이도·크기난이도를 상쇄해 존폭을 넓힌다.",
        ["구분", "등급/항목", "등급난이도", "기본존폭", "이동틱", "도주기본%", "도주증가%",
         "커서속도", "변속패턴", "스팟이동", "연속", "방향전환", "메모"],
        rows, {"메모": 26}, tab="1F3864",
    )

    rows = []
    for g in "EDCBASMLG":
        for rb in (0, 10, 20, 30, 40, 60):
            for er in (0, 30, 60):
                d = derive(g, 80, rb, er)
                if not d:
                    continue
                net, bar, zone, overflow, esc, inc = d
                rows.append([g, rb, er, net, bar, zone, round(zone / bar * 100, 1),
                             overflow or "", esc, inc, "", ])
    sheet(
        "29_난이도_도주_매트릭스", "장비 스탯이 실제로 사는 것 (80cm 물고기 기준)",
        "난이도 스탯(rodBonus)과 도망감소 스탯을 넣었을 때의 존폭·도주율. "
        "존폭비율 = 존폭/바폭 — 체감 성공률의 1차 지표. 초과난이도>0 이면 존폭이 이미 최소(1)로 눌린 상태다. "
        "★고등급은 난이도 스탯 없이는 존폭이 바닥이라 '등급 확률'만 올려도 실수령이 안 늘어난다.",
        ["등급", "난이도 스탯", "도망감소 스탯", "net", "바폭", "존폭", "존폭비율%",
         "초과난이도", "도주시작%", "미스당 도주+%", "메모"],
        rows, {"메모": 24}, tab="1F3864",
    )

    # ============================================================== 30 수리/마모
    em = read_java("parts/EquipmentManager.java")
    m = re.search(r"long gradeUnitRate\(String grade\) \{(.*?)\};", em, re.S)
    rate = {}
    if m:
        for g, v in re.findall(r'case "(\w)" -> (\d+)', m.group(1)):
            rate[g] = int(v)
        m2 = re.search(r"default -> (\d+);", m.group(1))
        if m2:
            rate["E"] = int(m2.group(1))
    if not rate:
        warn("gradeUnitRate 파싱 실패 — 수리비 시트 부실")
    CPH = 220  # 실측 처리량(캐스트/h) — balance-audit metrics 기준
    rows = []
    for t in ("릴", "줄", "바늘", "미끼", "찌"):
        for name, spec in parts.get(t, {}).items():
            pd = parse_part(spec)
            unit = rate.get(pd["등급"], rate.get("E", 5))
            rows.append([
                t, pd["이름"], pd["등급"], pd["내구"], unit,
                pd["내구"] * unit, unit, unit * CPH,
                round(pd["내구"] / CPH, 2), "",
            ])
    rows.sort(key=lambda r: (r[0], GRADE_RANK.get(r[2], 0), r[1]))
    sheet(
        "30_수리마모", "부품 내구 · 수리비 · 유지비",
        f"내구는 1캐스트에 1점 깎인다(낚싯대 본체의 '내구보존' 스탯 % 확률로 전체 스킵). "
        f"따라서 시간당 유지비는 내구 총량과 무관하게 '단가 × 캐스트/h' 다. 여기선 실측 {CPH}캐스트/h 가정. "
        "5부품 풀세팅이면 이 값의 5배가 시간당 고정 지출이다 — 설계 목표는 그 티어 수입의 10%.",
        ["부위", "이름", "등급", "최대내구", "내구1점 단가", "풀수리비(0→만)",
         "캐스트당 비용", f"시간당 유지비({CPH}캐스트/h)", "완전소모까지 시간(h)", "메모"],
        rows, {"메모": 24}, tab="375623",
        numfmt={"풀수리비(0→만)": "#,##0", f"시간당 유지비({CPH}캐스트/h)": "#,##0"},
    )

    # ============================================================== 31 조각 경제
    pf = read_java("parts/PartFragmentManager.java")
    m = re.search(r"int yieldForGrade\(String grade\) \{(.*?)\};", pf, re.S)
    yld = {g: int(v) for g, v in re.findall(r'case "(\w)" -> (\d+)', m.group(1))} if m else {}
    rows = [["부품 분해 수율", f"{g}등급", yld.get(g), "조각", "", ""] for g in "EDCBASMLG" if g in yld]
    TIER_LABEL = {0: "매우 흔함(12%↑) / 비드롭 중간재", 1: "흔함(8~11%)", 2: "보통(5~7%)", 3: "희귀(3~4%)", 4: "매우 희귀(1~2%)"}
    TIER_YIELD = {0: 1, 1: 1, 2: 2, 3: 3, 4: 4}
    TIER_COST = {0: 2, 1: 3, 2: 5, 3: 8, 4: 12}

    def rarity_tier(mid):
        c = max([d["chance"] for tbl in drops.values() for d in tbl if d["matId"] == mid] or [0])
        if c <= 0:
            return 0
        return 0 if c >= 12 else 1 if c >= 8 else 2 if c >= 5 else 3 if c >= 3 else 4

    for t in sorted(TIER_LABEL):
        rows.append(["재료 희귀도 tier", f"tier {t} — {TIER_LABEL[t]}", TIER_YIELD[t], "조각(분해)",
                     TIER_COST[t], "조각(제작비) — 항상 수율보다 커서 왕복 손실"])
    for mid, m2 in mats.items():
        t = rarity_tier(mid)
        bundle = 5 if m2.get("name") == "깨진 토기 조각" or mid == "깨진토기조각" else 1
        rows.append([
            "재료별", f"{m2.get('name', mid)}", TIER_YIELD[t], f"조각 / {bundle}개당",
            TIER_COST[t], f"tier {t}" + (" (묶음 5 — 유물감정 유입 과다로 하향)" if bundle > 1 else ""),
        ])
    sheet(
        "31_조각경제", "부품 조각 — 분해 수율 · 제작비",
        "조각은 부품/재료를 분해해 얻고 다른 재료를 만드는 데 쓴다. 제작비 > 분해수율 이라 왕복하면 항상 손실(싱크). "
        "희귀도 tier 는 '최대 드롭 확률'만 보므로 '흔한데 쓸 데도 없는' 재료를 구분 못 한다 — 묶음 단위가 그 보정 손잡이.",
        ["구분", "항목", "분해 수율", "단위", "제작비(조각)", "비고"], rows,
        {"항목": 32, "단위": 18, "비고": 44}, tab="375623",
    )

    # ============================================================== 90 점검표
    issues = []

    def add(sev, area, what, detail):
        issues.append([sev, area, what, detail, ""])

    # 스텁 지역인데 콘텐츠가 걸려 있음
    for rid in fishing_regions:
        if rid not in RG:
            add("높음", "지역", f"{rid}: regions.json 미등록", f"어종 {len(regionfish.get(rid, {}).get('기본', []))}종이 배정돼 있지만 갈 수 있는 지역이 아니다")
        elif is_stub(rid):
            add("높음", "지역", f"{rid}: 좌표 스텁", "pos1=pos2 — 지역 판정이 안 되므로 실제로 낚시 불가")
    for key, t in traps.items():
        rid = t.get("region", "")
        if rid not in RG or is_stub(rid):
            add("높음", "통발", f"{rid} 통발({t.get('variant')})", "설치 가능한 지역이 없는데 레시피는 판매 중")
    # 재료 획득/사용 불가
    for mid, m in mats.items():
        if not mat_paths.get(mid):
            add("높음", "재료", f"{m.get('name', mid)}: 획득 경로 없음",
                "드롭테이블·JSON 레시피·코드(드릴/보스/RecipeLoader) 어디에도 산출처가 없다")
        if mid not in mat_use:
            add("낮음", "재료", f"{m.get('name', mid)}: 사용처 없음",
                f"어떤 레시피도 이 재료를 쓰지 않는다 (획득경로: {' / '.join(mat_paths.get(mid, [])) or '없음'})")
    for rid, rc in recipes.items():
        for i in rc.get("ingredients", []):
            if i.get("kind") == "custom" and i["typeOrMatId"] not in mats:
                if not i["typeOrMatId"].startswith(("작물_", "채집_", "강화")):
                    add("보통", "레시피", f"{rid} 재료 미정의", f"{i['typeOrMatId']} 가 materials.json 에 없다")
    # 레시피 없는 장비
    for t, items in parts.items():
        for name, spec in items.items():
            pd = parse_part(spec)
            has = (rec_by_rod.get(name) if t == "낚싯대" else rec_by_part.get((t, name))) or rec_by_part.get((t, name))
            if not has and pd["가격"] > 0:
                # 잠수상점·튜토 등은 애초에 '돈으로만' 파는 경로라 정상일 수 있다 → 심각도 낮춤
                money_only = any(k in pd["출처"] for k in ("잠수상점", "튜토", "대장간", "심해"))
                add("낮음" if money_only else "보통", "장비", f"{t} {name}: 레시피 없음",
                    f"상점가 {pd['가격']:,}원 / 출처 {pd['출처']} — 돈으로만 사는 항목이면 정상, "
                    "사다리 중간이면 재료 소모처가 빠진 것")
            if not has and pd["가격"] == 0 and "히든" in pd["출처"]:
                add("높음", "장비", f"{t} {name}: 획득 불가", "가격 0 + 레시피 없음 + 히든 출처")
    # 등급 사다리 가격 역전
    for t, items in parts.items():
        by_grade = defaultdict(list)
        for name, spec in items.items():
            pd = parse_part(spec)
            by_grade[pd["등급"]].append(pd)
        prev_max = None
        for g in "EDCBASMLG":
            lst = [x for x in by_grade.get(g, []) if x["가격"] > 0]
            if not lst:
                continue
            mn = min(x["가격"] for x in lst)
            if prev_max is not None and mn < prev_max * 0.5:
                add("보통", "가격", f"{t} {g}등급 가격 역전", f"{g}등급 최저가 {mn:,} < 하위등급 최고가 {prev_max:,} 의 절반")
            prev_max = max(x["가격"] for x in lst)
    # 지역별 고등급 병목
    for rid in fishing_regions:
        if rid not in RG or is_stub(rid):
            continue
        bk = grade_buckets(pool_for(rid))
        for g in "SMLG":
            if len(bk.get(g, [])) == 1:
                add("낮음", "어종풀", f"{rid} {g}등급 어종 1종뿐", f"{bk[g][0]} — 그 등급이 나오면 100% 이 어종")
        missing = [g for g in "BASMLG" if not bk.get(g)]
        if missing:
            add("낮음", "어종풀", f"{rid}: {''.join(missing)}등급 없음", "해당 등급 롤은 건너뛰어지고 피티가 보존된다")
        if not bk.get("E"):
            up = next((g for g in "DCBASMLG" if bk.get(g)), None)
            if up:
                add("보통", "어종풀", f"{rid}: E등급 어종 없음 → 실패분이 전부 {up}등급으로 승격",
                    f"롤 실패(=E)가 가용성 폴백으로 {up}로 올라간다. 이 지역 하한 수입이 조용히 크게 뛴다")
    # 드롭테이블 없는 낚시 지역
    for rid in fishing_regions:
        if rid not in drops and not any(p in drops for p in parent_chain(rid)):
            add("보통", "재료", f"{rid}: 재료 드롭테이블 없음", "이 지역에서 낚시해도 조합 재료가 전혀 안 나온다")
    # 요리 재료 미해결
    for did, d in dishes.items():
        for i in d.get("ingredients", []):
            if i.get("kind") == "custom" and i["typeOrMatId"] not in mats:
                if not any(i["typeOrMatId"].startswith(pfx) for pfx in ("작물_", "채집_", "강화")):
                    add("보통", "요리", f"{did}: 재료 {i['typeOrMatId']} 미정의", "materials.json/작물/채집 어디서도 해석 안 됨")
    # 상점 회수율
    for c in SH.get("categories", []):
        if isinstance(c, dict) and c.get("buy") and c.get("sell") and c["sell"] >= c["buy"]:
            add("높음", "경제", f"상점 {c.get('key')} 매도≥매입", f"매입 {c['buy']} / 매도 {c['sell']} — 무한 차익")

    sev_rank = {"높음": 0, "보통": 1, "낮음": 2}
    issues.sort(key=lambda r: (sev_rank.get(r[0], 9), r[1], r[2]))
    ws = sheet(
        "90_점검표", "자동 점검 결과 — 여기서부터 보면 빠르다",
        "기계가 데이터 정합성만 본 것이다(수치가 재미있는지는 판단하지 않는다). "
        "'높음'은 콘텐츠가 실제로 작동하지 않을 가능성이 있는 항목.",
        ["심각도", "영역", "항목", "설명", "판단/메모"], issues,
        {"항목": 40, "설명": 60, "판단/메모": 26}, tab="C00000",
    )
    for r in range(5, 5 + len(issues)):
        sev = ws.cell(row=r, column=1).value
        if sev == "높음":
            for c in range(1, 6):
                ws.cell(row=r, column=c).fill = BAD_FILL
        elif sev == "보통":
            for c in range(1, 6):
                ws.cell(row=r, column=c).fill = WARN_FILL

    # ============================================================== 99 목차 (00 시트 뒤에)
    ws = wb.create_sheet("99_시트목차")
    ws["A1"] = "시트 목차"
    ws["A1"].font = TITLE_FONT
    for i, h in enumerate(["시트", "내용", "행수"], 1):
        c = ws.cell(row=3, column=i, value=h)
        c.fill = HDR_FILL
        c.font = HDR_FONT
    for r, (n, t, cnt) in enumerate(index_rows, 4):
        ws.cell(row=r, column=1, value=n)
        ws.cell(row=r, column=2, value=t)
        ws.cell(row=r, column=3, value=cnt)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 56
    ws.column_dimensions["C"].width = 10
    if WARNINGS:
        r0 = len(index_rows) + 6
        ws.cell(row=r0, column=1, value="생성 시 경고 (수치 신뢰도에 영향)").font = Font(bold=True, color="C00000")
        for j, w in enumerate(WARNINGS, r0 + 1):
            ws.cell(row=j, column=1, value=w)

    # 시트 순서: 00 읽는법 → 99 목차 → 나머지
    order = wb.sheetnames
    wb.move_sheet("00_읽는법", offset=-order.index("00_읽는법"))
    wb.move_sheet("99_시트목차", offset=-(wb.sheetnames.index("99_시트목차") - 1))

    wb.save(out_path)
    return out_path


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", default=os.path.expanduser("~/Desktop/바르칸_밸런스_통합.xlsx"))
    ap.add_argument("--json", help="BlockShip 데이터 디렉터리 override")
    ap.add_argument("--java", help="blockship-plugin src/main/java/com/blockship override")
    ap.add_argument("--sim", type=int, default=400000, help="등급 PRD 몬테카를로 캐스트 수")
    a = ap.parse_args()
    global JSON_ROOT, JAVA_ROOT
    if a.json:
        JSON_ROOT = a.json
    if a.java:
        JAVA_ROOT = a.java
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        print("openpyxl 필요:  python3 -m pip install --user openpyxl", file=sys.stderr)
        return 1
    path = build_workbook(a.out, a.sim)
    print(f"\n생성 완료: {path}")
    if WARNINGS:
        print(f"경고 {len(WARNINGS)}건 — 99_시트목차 시트 하단 참조")
    return 0


if __name__ == "__main__":
    sys.exit(main())
